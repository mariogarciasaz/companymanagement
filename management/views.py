from typing import Any, Dict
from django.db.models.query import QuerySet
from django.forms.models import BaseModelForm
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from management.models import Employee, Product, Invoice, Category
from clients.models import Client
from management.forms import EmployeeForm, ProductForm, InvoiceForm, CategoryForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.db.models import Q
from django.urls import reverse_lazy
import datetime
import os
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
from .pdf import html_to_pdf
from django.db.models.functions import ExtractYear
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import JsonResponse

# Create your views here.


#EMPLOYEE VIEWS
class EmployeeListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Employee
    template_name = 'employees_list.html'
    context_object_name = 'employees'
    login_url = 'website:login'
    permission_required = 'management.view_employee'


    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(category__icontains=search_query)
            )
        return queryset
    


class AddEmployee(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'add_employee.html'
    success_url = reverse_lazy('management:employees')
    login_url = 'website:login'
    permission_required = 'management.add_employee'

    def form_valid(self, form):
        form.instance.name = self.request.POST.get('name')
        return super().form_valid(form)


class UpdateEmployee(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'update_employee.html'
    success_url = reverse_lazy('management:employees')
    login_url = 'website:login'
    permission_required = 'management.change_employee'

    def form_valid(self, form):
        form.instance.employee = self.request.user.employee
        return super().form_valid(form)
    


class DeleteEmployee(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Employee
    template_name = 'employee_confirm_delete.html'
    success_url = reverse_lazy('management:employees')
    login_url = 'website:login'
    permission_required = 'management.delete_employee'



#PRODUCT VIEWS

class ProductListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Product
    template_name = 'products_list.html'
    context_object_name = 'products'
    login_url = 'website:login'
    permission_required = 'management.view_product'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = Product.objects.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(category__name__icontains=search_query)
            )
        return queryset
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        product_list = Product.objects.all()

        total_products = len(product_list)
        out_stock = []
        out_stock2 = []

        for product in product_list:
            if product.quantity <= 0:
                out_stock.append(product)
            elif product.quantity >= 1 and product.quantity < 10:
                out_stock2.append(product)

        total_out_stock = len(out_stock)
        total_out_stock2 = len(out_stock2)
        context['total_out_stock'] = total_out_stock
        context['total_out_stock2'] = total_out_stock2
        context['total_products'] = total_products

        return context
    

class AddProduct(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'add_product.html'
    success_url = reverse_lazy('management:products')
    login_url = 'website:login'
    permission_required = 'management.add_product'

    def form_valid(self, form):
        form.instance.name = self.request.POST.get('name')
        return super().form_valid(form)
    

class UpdateProduct(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'update_product.html'
    success_url = reverse_lazy('management:products')
    login_url = 'website:login'
    permission_required = 'management.change_product'

    def form_valid(self, form):
        form.instance.name = self.request.POST.get('name')
        return super().form_valid(form)


class DeleteProduct(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Product
    template_name = 'product_confirm_delete.html'
    success_url = reverse_lazy('management:products')
    login_url = 'website:login'
    permission_required = 'management.delete_product'


#CATEGORY VIEWS


class CategoryListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Category
    template_name = 'categories_list.html'
    context_object_name = 'categories'
    login_url = 'website:login'
    permission_required = 'management.view_category'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = Product.objects.filter(
                Q(name__icontains=search_query)
            )
        return queryset
    

class AddCategory(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'add_category.html'
    success_url = reverse_lazy('management:categories')
    login_url = 'website:login'
    permission_required = 'management.add_category'

    def form_valid(self, form):
        form.instance.name = self.request.POST.get('name')
        return super().form_valid(form)
    

class UpdateCategory(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'update_category.html'
    success_url = reverse_lazy('management:categories')
    login_url = 'website:login'
    permission_required = 'management.change_category'

    def form_valid(self, form):
        form.instance.name = self.request.POST.get('name')
        return super().form_valid(form)
    

class DeleteCategory(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Category
    template_name = 'category_confirm_delete.html'
    success_url = reverse_lazy('management:categories')
    login_url = 'website:login'
    permission_required = 'management.delete_category'



#INVOICE VIEWS

class InvoiceListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Invoice
    template_name = 'invoices_list.html'
    context_object_name = 'invoices'
    login_url = 'website:login'
    permission_required = 'management.view_invoice'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        year = self.request.GET.get('year')
        if search_query:
            queryset = Invoice.objects.filter(
                Q(client__full_name__icontains=search_query) |
                Q(client_car__icontains=search_query) |
                Q(client_car_model__icontains=search_query) |
                Q(invoice_number__icontains=search_query)|
                Q(state__icontains=search_query)
            )

        if year:
            # Filtrar por el año seleccionado
            queryset = queryset.filter(invoice_date__year=year)

        return queryset
    
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        invoice_list = Invoice.objects.all()

        for invoice in invoice_list:
            if invoice.invoice_due_date <= datetime.date.today() and invoice.paid == False:
                invoice.state = "Expirada"
                invoice.save()
            elif invoice.invoice_due_date <= datetime.date.today() and invoice.paid == True:
                invoice.state = "Pagada"
                invoice.save()
            elif invoice.paid == True:
                invoice.state = "Pagada"
                invoice.save()
            elif invoice.paid == True and invoice.invoice_due_date > datetime.date.today():
                invoice.state = "Pagada"
                invoice.save()
            elif invoice.paid == False:
                invoice.state = "Pendiente"
                invoice.save()

        total_invoices = len(invoice_list)
        expired_invoices = len (invoice_list.filter(state="Expirada"))
        paid_invoices = len (invoice_list.filter(state="Pagada"))
        pending_invoices = len (invoice_list.filter(state="Pendiente"))



        context['total_invoices'] = total_invoices
        context['expired_invoices'] = expired_invoices
        context['paid_invoices'] = paid_invoices
        context['pending_invoices'] = pending_invoices
        


        # Obtener los años únicos de las facturas
        years = Invoice.objects.dates('invoice_date', 'year', order='DESC')
        years = [year.year for year in years]
        context['years'] = years
        
        # Obtener el año seleccionado en el filtro
        selected_year = self.request.GET.get('year')
        context['selected_year'] = selected_year

        return context
    

    def generate_excel_report(request):


        # The following code is written in Python and is used to filter the issued invoices based on the
        # year provided in the GET request. It retrieves the year value from the GET request using the
        # `request.GET.get('year')` method and then filters the `issuedinvoices` queryset based on the
        # year using the `filter()` method. The `date__year` parameter in the `filter()` method is
        # used to filter the queryset based on the year value.
        year = request.GET.get('year')
        queryset = Invoice.objects.filter(invoice_date__year=year).order_by('id')


        # The following code is creating a new Excel workbook and activating the first sheet in the
        # workbook.
        workbook = Workbook()
        sheet = workbook.active


        # The following code is formatting an Excel sheet by merging cells from row 1 to row 3 and columns
        # 1 to 12, adding a title "Report of Issued Invoices" in the merged cell, centering the text,
        # setting the font to bold, size 20, and color white, and filling the cell with a blue color.
        # It also sets the font of cell A1 to bold, size 20, color white, and italic. Finally, it sets
        # the height of row 5 to 30.
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=9)
        merged_cell = sheet.cell(row=1, column=1)
        merged_cell.value = "Reporte de Facturas"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True, size=20, color='FFFFFF')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell.fill = fill
        sheet['A1'].font = Font(bold=True, size=20, color='FFFFFF', italic=True)
        row_dimension = sheet.row_dimensions[5]
        row_dimension.height = 30


        # The following code is setting the width of all columns in an Excel sheet to 20. The variable
        # `all_columns` contains a list of all the column names, and the `for` loop iterates over each
        # column and sets its width to 20 using the `column_dimensions` attribute of the sheet object.
        all_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P']
        for column in all_columns:
            sheet.column_dimensions[column].width = 20


        # The following code is defining a list of headers for a table or spreadsheet. The headers are in
        # multiple languages, as indicated by the use of the `_()` function to translate them. The
        # code then adds three empty rows and the headers to the sheet.
        headers = ['Client', 'Factura', 'Fecha Factura', 'Fecha Expiracion', 'Subtotal', 'VAT', 'Descuento', 'Total', 'Estado']
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append(headers)


        # The following code is defining a border style for a header in Python using the `Border` class
        # from the `openpyxl.styles` module. The border style includes thin borders on all sides (top,
        # bottom, left, and right) of the header.
        header_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )


        # The following code is formatting the header cells of a spreadsheet. It sets the fill color to a
        # light blue shade, makes the font bold and white with a size of 14, centers the text
        # horizontally and vertically, and applies a border to the cells. The code is using the
        # openpyxl library to access and modify the spreadsheet.
        for header_cell in sheet[5]:
            header_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            header_cell.font = Font(bold=True, color='FFFFFF', size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = header_border


        # The following code is iterating over a queryset of invoices and appending the data of each
        # invoice to a row in a sheet. The row is then added to the sheet and the row index is
        # incremented. The data includes the client name, CIF, invoice number, date, taxable base, IVA
        # rate, IVA, retention, total, due date, paid status, and state of the invoice.
        row_index = 3
        for invoice in queryset:
            row = [
                f"{invoice.client.name + ' ' + invoice.client.last_name}",
                invoice.invoice_number,
                invoice.invoice_date,
                invoice.invoice_due_date,
                f"{invoice.subtotal} $",
                f"{invoice.vat} %",
                f"{invoice.discount} %",
                f"{invoice.total} $",
                invoice.state
            ]
            sheet.append(row)
            row_index += 1


        # The following code is defining a border style for a cell in Python using the `Border` class from
        # the `openpyxl.styles` module. The border style includes thin borders on all four sides of
        # the cell (top, bottom, left, and right).
        data_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )


        # The following code is iterating through all the rows in a sheet (starting from row 6 and ending
        # at row 9000) and then iterating through all the cells in each row. For each cell, it is
        # setting the alignment to be centered both horizontally and vertically, and setting the
        # border to be a specific style (which is defined elsewhere in the code as "data_border").
        for row in sheet.iter_rows(min_row=6, max_row=row_index+2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = data_border
        


        # TOTAL CHARGED

        charged = []
        for invoice in queryset:
            charged.append(invoice.total)

        total_charged = sum(charged)

        sheet.merge_cells(start_row=5, start_column=11, end_row=5, end_column=13)
        merged_cell_total_charged_title = sheet.cell(row=5, column=11)
        merged_cell_total_charged_title.value = "TOTAL FACTURADO"
        merged_cell_total_charged_title.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged_title.font = Font(bold=True, size=15, color='FFFFFF')

        sheet.merge_cells(start_row=6, start_column=11, end_row=6, end_column=13)
        merged_cell_total_charged = sheet.cell(row=6, column=11)
        merged_cell_total_charged.value = f"{total_charged} $"
        merged_cell_total_charged.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged.font = Font(bold=True, size=15, color='000000')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell_total_charged_title.fill = fill
        merged_cell_total_charged.column = 11

        for row in sheet.iter_rows(min_row=5, max_row=6, min_col=11, max_col=13):
            for cell in row:
                cell.border = data_border

        filename = f"Reporte Facturas {year}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


    def print_pdf(request):
        # Render the 'invoice_pdf.html' template with the provided context data (if any)
        rendered_template = render(request, "invoice_pdf.html")
    
        # Create an HttpResponse object with the rendered content and return it
        return HttpResponse(rendered_template)




class ViewInvoice(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    model = Invoice
    form_class = InvoiceForm
    template_name ='invoice_detail.html'
    login_url = 'website:login'
    permission_required='management.view_invoice',

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AddInvoice(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = 'add_invoice.html'
    success_url = reverse_lazy('management:invoices')
    login_url = 'website:login'
    permission_required='management.add_invoice',

    def form_valid(self, form):
        invoice = form.save(commit=False)

        
        line_one_quantity = form.cleaned_data.get('line_one_quantity')
        line_two_quantity = form.cleaned_data.get('line_two_quantity')
        line_three_quantity = form.cleaned_data.get('line_three_quantity')
        line_four_quantity = form.cleaned_data.get('line_four_quantity')
        line_five_quantity = form.cleaned_data.get('line_five_quantity')
        line_six_quantity = form.cleaned_data.get('line_six_quantity')
        line_seven_quantity = form.cleaned_data.get('line_seven_quantity')
        line_eight_quantity = form.cleaned_data.get('line_eight_quantity')
        line_nine_quantity = form.cleaned_data.get('line_nine_quantity')
        line_ten_quantity = form.cleaned_data.get('line_ten_quantity')

        if line_one_quantity:
            product_line_1 = form.cleaned_data.get('line_one')
            product_line_1.quantity -= line_one_quantity
            product_line_1.save()

        if line_two_quantity:
            product_line_2 = form.cleaned_data.get('line_two')
            product_line_2.quantity -= line_two_quantity
            product_line_2.save()

        if line_three_quantity:
            product_line_3 = form.cleaned_data.get('line_three')
            product_line_3.quantity -= line_three_quantity
            product_line_3.save()

        if line_four_quantity:
            product_line_4 = form.cleaned_data.get('line_four')
            product_line_4.quantity -= line_four_quantity
            product_line_4.save()

        if line_five_quantity:
            product_line_5 = form.cleaned_data.get('line_five')
            product_line_5.quantity -= line_five_quantity
            product_line_5.save()
            invoice.save()
        
        if line_six_quantity:
            product_line_6 = form.cleaned_data.get('line_six')
            product_line_6.quantity -= line_six_quantity
            product_line_6.save()
            invoice.save()

        if line_seven_quantity:
            product_line_7 = form.cleaned_data.get('line_seven')
            product_line_7.quantity -= line_seven_quantity
            product_line_7.save()
            invoice.save()

        if line_eight_quantity:
            product_line_8 = form.cleaned_data.get('line_eight')
            product_line_8.quantity -= line_eight_quantity
            product_line_8.save()
            invoice.save()

        if line_nine_quantity:
            product_line_9 = form.cleaned_data.get('line_nine')
            product_line_9.quantity -= line_nine_quantity
            product_line_9.save()
            invoice.save()

        if line_ten_quantity:
            product_line_10 = form.cleaned_data.get('line_ten')
            product_line_10.quantity -= line_ten_quantity
            product_line_10.save()
            invoice.save()

        return super().form_valid(form)
    



def get_product_price(request, product_id):
    product = Product.objects.get(id=product_id)
    return JsonResponse({'unit_price': product.price})


def get_client_datas(request, client_id):
    client = Client.objects.get(id=client_id)
    return JsonResponse({'client_address': client.address
                        , 'client_phone': client.phone
                        , 'client_email': client.email
                        })




class UpdateInvoice(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = 'update_invoice.html'
    success_url = reverse_lazy('management:invoices')
    login_url = 'website:login'
    permission_required='management.change_invoice',

    def form_valid(self, form):
        # Guardar la instancia de la factura sin guardarla en la base de datos
        invoice = form.save(commit=False)

        # Obtener la factura original de la base de datos
        original_invoice = Invoice.objects.get(pk=invoice.pk)

        # Restaurar la cantidad original de productos en el inventario
        self.restore_product_quantity(original_invoice)

        # Restar o sumar la diferencia de cantidad de productos al inventario
        self.update_product_quantity(original_invoice, invoice)

        # Guardar la instancia de la factura en la base de datos
        invoice.save()

        return super().form_valid(form)

    def restore_product_quantity(self, invoice):
        # Restaurar la cantidad original de productos en el inventario
        for line_num in ['one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten']:
            description_field = f'line_{line_num}'
            quantity_field = f'line_{line_num}_quantity'
            description = getattr(invoice, description_field)
            quantity = getattr(invoice, quantity_field)
            if description:
                description.quantity += quantity
                description.save()

    def update_product_quantity(self, original_invoice, updated_invoice):
        # Restar o sumar la diferencia de cantidad de productos al inventario
        for line_num in ['one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten']:
            description_field = f'line_{line_num}'
            quantity_field = f'line_{line_num}_quantity'
            original_description = getattr(original_invoice, description_field)
            original_quantity = getattr(original_invoice, quantity_field)
            updated_description = getattr(updated_invoice, description_field)
            updated_quantity = getattr(updated_invoice, quantity_field)

            if original_description and updated_description:
                quantity_diff = updated_quantity - original_quantity

                if quantity_diff > 0:
                    # La cantidad actual es mayor que la original, restar la diferencia del stock
                    updated_description.quantity -= quantity_diff
                elif quantity_diff < 0:
                    # La cantidad actual es menor que la original, sumar la diferencia al stock
                    updated_description.quantity += abs(quantity_diff)

                updated_description.save()



class DeleteInvoice(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Invoice
    template_name = 'invoice_confirm_delete.html'
    success_url = reverse_lazy('management:invoices')
    login_url = 'website:login'
    permission_required='management.delete_invoice',

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        restore_stock = self.object.restore_stock

        if restore_stock:
            # Sumar el stock antes de eliminar la factura
            self.restore_product_quantity(self.object)

        # Eliminar la factura
        success_url = self.get_success_url()
        self.object.delete()

        return HttpResponseRedirect(success_url)

    def restore_product_quantity(self, invoice):
        # Sumar la cantidad de productos de la factura al inventario

        for line_num in range(1, 6):
            description_field = f'line_{line_num}_description'
            quantity_field = f'line_{line_num}_quantity'
            description = getattr(invoice, description_field)
            quantity = getattr(invoice, quantity_field)

            if description:
                description.quantity += quantity
                description.save()





