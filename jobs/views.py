from typing import Any, Dict
from django.db.models.query import QuerySet
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import render
from jobs.models import Job
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db.models import Q
from django.urls import reverse_lazy
from jobs.forms import JobForm
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string




class JobListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Job
    template_name = 'jobs_list.html'
    context_object_name = 'jobs'
    login_url = 'website:login'
    permission_required = 'jobs.view_job'



    def get_queryset(self):
        queryset=super().get_queryset()
        search_query = self.request.GET.get('search')
        year = self.request.GET.get('year')
        if search_query:
            queryset= Job.objects.filter(
                Q(car__icontains=search_query) |
                Q(car_model__icontains=search_query) |
                Q(employee__name__icontains=search_query) |
                Q(state__icontains=search_query)
            )

        if year:
            queryset = queryset.filter(start_date__year=year)

        return queryset
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        jobs_list = Job.objects.all()


        for job in jobs_list:
            if job.finished == False:
                job.state = 'En Curso'
                job.save()
            elif job.finished == True:
                job.state = 'Finalizado'
                job.save()
    

        total_jobs = len(jobs_list)
        finished_jobs = len(jobs_list.filter(finished=True))
        unfinished_jobs = len(jobs_list.filter(finished=False))

        context['total_jobs'] = total_jobs
        context['finished_jobs'] = finished_jobs
        context['unfinished_jobs'] = unfinished_jobs

        years = Job.objects.dates('start_date', 'year', order='DESC')
        years = [year.year for year in years]
        context['years'] = years

        selected_year = self.request.GET.get('year')
        context['selected_year'] = selected_year

        return context
    

    def generate_excel_report(request):


        # The following code is written in Python and is used to filter the issued invoices based on the
        # year provided in the GET request. It retrieves the year value from the GET request using the
        # `request.GET.get('year')` method and then filters the `issuedinvoices` queryset based on the
        # year using the `filter()` method. The `date__year` parameter in the `filter()` method is
        # used to filter the queryset based on the year value.
        year = request.GET.get('year')
        queryset = Job.objects.filter(start_date__year=year).order_by('id')


        # The following code is creating a new Excel workbook and activating the first sheet in the
        # workbook.
        workbook = Workbook()
        sheet = workbook.active


        # The following code is formatting an Excel sheet by merging cells from row 1 to row 3 and columns
        # 1 to 12, adding a title "Report of Issued Invoices" in the merged cell, centering the text,
        # setting the font to bold, size 20, and color white, and filling the cell with a blue color.
        # It also sets the font of cell A1 to bold, size 20, color white, and italic. Finally, it sets
        # the height of row 5 to 30.
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=7)
        merged_cell = sheet.cell(row=1, column=1)
        merged_cell.value = "Reporte de Trabajos"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True, size=20, color='FFFFFF')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell.fill = fill
        sheet['A1'].font = Font(bold=True, size=20, color='FFFFFF', italic=True)
        row_dimension = sheet.row_dimensions[5]
        row_dimension.height = 30


        # The following code is setting the width of all columns in an Excel sheet to 20. The variable
        # `all_columns` contains a list of all the column names, and the `for` loop iterates over each
        # column and sets its width to 20 using the `column_dimensions` attribute of the sheet object.
        all_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P']
        for column in all_columns:
            sheet.column_dimensions[column].width = 20


        # The following code is defining a list of headers for a table or spreadsheet. The headers are in
        # multiple languages, as indicated by the use of the `_()` function to translate them. The
        # code then adds three empty rows and the headers to the sheet.
        headers = ['Client', 'Coche', 'Modelo', 'Empleado', 'Descripcion', 'Fecha Inicio', 'Fecha Fin']
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append(headers)


        # The following code is defining a border style for a header in Python using the `Border` class
        # from the `openpyxl.styles` module. The border style includes thin borders on all sides (top,
        # bottom, left, and right) of the header.
        header_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )


        # The following code is formatting the header cells of a spreadsheet. It sets the fill color to a
        # light blue shade, makes the font bold and white with a size of 14, centers the text
        # horizontally and vertically, and applies a border to the cells. The code is using the
        # openpyxl library to access and modify the spreadsheet.
        for header_cell in sheet[5]:
            header_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            header_cell.font = Font(bold=True, color='FFFFFF', size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = header_border


        # The following code is iterating over a queryset of invoices and appending the data of each
        # invoice to a row in a sheet. The row is then added to the sheet and the row index is
        # incremented. The data includes the client name, CIF, invoice number, date, taxable base, IVA
        # rate, IVA, retention, total, due date, paid status, and state of the invoice.
        row_index = 3
        for job in queryset:
            row = [
                f"{job.client.name + ' ' + job.client.last_name}",
                job.car,
                job.car_model,
                f"{job.employee.name + ' ' + job.employee.last_name}" ,
                job.description,
                job.start_date,
                job.end_date,
            ]
            sheet.append(row)
            row_index += 1


        # The following code is defining a border style for a cell in Python using the `Border` class from
        # the `openpyxl.styles` module. The border style includes thin borders on all four sides of
        # the cell (top, bottom, left, and right).
        data_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )


        # The following code is iterating through all the rows in a sheet (starting from row 6 and ending
        # at row 9000) and then iterating through all the cells in each row. For each cell, it is
        # setting the alignment to be centered both horizontally and vertically, and setting the
        # border to be a specific style (which is defined elsewhere in the code as "data_border").
        for row in sheet.iter_rows(min_row=6, max_row=row_index+2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = data_border
        


        # TOTAL JOBS

        jobs = []
        for job in queryset:
            jobs.append(job.client)

        total_jobs = len(jobs)

        sheet.merge_cells(start_row=5, start_column=9, end_row=5, end_column=11)
        merged_cell_total_charged_title = sheet.cell(row=5, column=9)
        merged_cell_total_charged_title.value = "TOTAL TRABAJOS"
        merged_cell_total_charged_title.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged_title.font = Font(bold=True, size=15, color='FFFFFF')

        sheet.merge_cells(start_row=6, start_column=9, end_row=6, end_column=11)
        merged_cell_total_charged = sheet.cell(row=6, column=9)
        merged_cell_total_charged.value = f"{total_jobs}"
        merged_cell_total_charged.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged.font = Font(bold=True, size=15, color='000000')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell_total_charged_title.fill = fill
        merged_cell_total_charged.column = 9

        for row in sheet.iter_rows(min_row=5, max_row=6, min_col=9, max_col=11):
            for cell in row:
                cell.border = data_border

        filename = f"Reporte Trabajos {year}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response





class AddJob(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Job
    form_class = JobForm
    template_name = 'add_job.html'
    success_url = reverse_lazy('jobs:jobs')
    login_url = 'website:login'
    permission_required = 'jobs.add_job'
    success_message = "Trabajo registrado con exito"

    def form_valid(self, form):
        form.instance.car = self.request.POST.get('car')
        return super().form_valid(form)
    


class UpdateJob(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Job
    form_class = JobForm
    template_name = 'edit_job.html'
    success_url = reverse_lazy('jobs:jobs')
    login_url = 'website:login'
    permission_required = 'jobs.change_job'

    def form_valid(self, form):
        form.instance.car = self.request.POST.get('car')
        return super().form_valid(form)



class DeleteJob(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Job
    template_name = 'job_confirm_delete.html'
    success_url = reverse_lazy('jobs:jobs')
    login_url = 'website:login'
    permission_required = 'jobs.delete_job'



